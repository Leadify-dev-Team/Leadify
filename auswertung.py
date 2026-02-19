import flet as ft
from datetime import datetime
from database import Database
import csv
from pathlib import Path
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


# ============================================================================
# MANAGER-KLASSE (Business-Logik / Datenbankoperationen)
# ============================================================================

class AuswertungManager:
    """Verwaltet Operationen für die Auswertung aller Leads"""
    
    def __init__(self, db: Database):
        self.db = db
    
    def get_all_leads(self):
        """Gibt ALLE Leads zurück (unabhängig von Status und Bearbeiter)"""
        sql = """
            SELECT 
                l.*,
                CONCAT(b_erfasser.vorname, ' ', b_erfasser.nachname) as erfasser_name,
                CONCAT(b_bearbeiter.vorname, ' ', b_bearbeiter.nachname) as bearbeiter_name,
                f.name as kunde_name,
                CONCAT(ap.vorname, ' ', ap.nachname) as ansprechpartner_name,
                ap.email as kunde_email,
                ap.telefon as kunde_telefon,
                pos.bezeichnung as ansprechpartner_position,
                p.produkt as produkt_name,
                pg.produkt as produktgruppe_name,
                s.status as status_name,
                pz.zustand as produktzustand_name,
                q.quelle as quelle_name
            FROM lead l
            LEFT JOIN benutzer b_erfasser ON l.erfasser_id = b_erfasser.benutzer_id
            LEFT JOIN benutzer b_bearbeiter ON l.bearbeiter_id = b_bearbeiter.benutzer_id
            LEFT JOIN ansprechpartner ap ON l.ansprechpartner_id = ap.id
            LEFT JOIN firma f ON ap.firma_id = f.id
            LEFT JOIN position pos ON ap.position_id = pos.id
            LEFT JOIN produkte p ON l.produkt_id = p.produkt_id
            LEFT JOIN produktgruppe pg ON l.produktgruppe_id = pg.produkt_id
            LEFT JOIN status s ON l.status_id = s.id
            LEFT JOIN produktzustand pz ON l.produktzustand_id = pz.id
            LEFT JOIN quelle q ON l.quelle_id = q.id_quelle
            ORDER BY l.datum_erfasst DESC
        """
        return self.db.fetch_all(sql)
    
    def get_leads_by_bearbeiter(self, bearbeiter_id: int):
        """Gibt alle Leads eines bestimmten Bearbeiters zurück"""
        sql = """
            SELECT 
                l.*,
                CONCAT(b_erfasser.vorname, ' ', b_erfasser.nachname) as erfasser_name,
                CONCAT(b_bearbeiter.vorname, ' ', b_bearbeiter.nachname) as bearbeiter_name,
                f.name as kunde_name,
                CONCAT(ap.vorname, ' ', ap.nachname) as ansprechpartner_name,
                p.produkt as produkt_name,
                s.status as status_name
            FROM lead l
            LEFT JOIN benutzer b_erfasser ON l.erfasser_id = b_erfasser.benutzer_id
            LEFT JOIN benutzer b_bearbeiter ON l.bearbeiter_id = b_bearbeiter.benutzer_id
            LEFT JOIN ansprechpartner ap ON l.ansprechpartner_id = ap.id
            LEFT JOIN firma f ON ap.firma_id = f.id
            LEFT JOIN produkte p ON l.produkt_id = p.produkt_id
            LEFT JOIN status s ON l.status_id = s.id
            WHERE l.bearbeiter_id = ?
            ORDER BY l.datum_erfasst DESC
        """
        return self.db.fetch_all(sql, (bearbeiter_id,))
    
    def get_leads_by_status(self, status_id: int):
        """Gibt alle Leads mit einem bestimmten Status zurück"""
        sql = """
            SELECT 
                l.*,
                CONCAT(b_erfasser.vorname, ' ', b_erfasser.nachname) as erfasser_name,
                CONCAT(b_bearbeiter.vorname, ' ', b_bearbeiter.nachname) as bearbeiter_name,
                f.name as kunde_name,
                p.produkt as produkt_name,
                s.status as status_name
            FROM lead l
            LEFT JOIN benutzer b_erfasser ON l.erfasser_id = b_erfasser.benutzer_id
            LEFT JOIN benutzer b_bearbeiter ON l.bearbeiter_id = b_bearbeiter.benutzer_id
            LEFT JOIN ansprechpartner ap ON l.ansprechpartner_id = ap.id
            LEFT JOIN firma f ON ap.firma_id = f.id
            LEFT JOIN produkte p ON l.produkt_id = p.produkt_id
            LEFT JOIN status s ON l.status_id = s.id
            WHERE l.status_id = ?
            ORDER BY l.datum_erfasst DESC
        """
        return self.db.fetch_all(sql, (status_id,))
    
    def get_all_erfasser(self):
        """Gibt alle Erfasser zurück (für Filter)"""
        sql = """
            SELECT DISTINCT b.benutzer_id, CONCAT(b.vorname, ' ', b.nachname) as name
            FROM benutzer b
            INNER JOIN lead l ON l.erfasser_id = b.benutzer_id
            WHERE b.is_approved = 1
            ORDER BY b.vorname, b.nachname
        """
        return self.db.fetch_all(sql)
    
    def get_statistics(self):
        """Gibt Statistiken über alle Leads zurück"""
        sql = """
            SELECT 
                COUNT(*) as gesamt,
                SUM(CASE WHEN status_id = 1 THEN 1 ELSE 0 END) as offen,
                SUM(CASE WHEN status_id = 2 THEN 1 ELSE 0 END) as in_bearbeitung,
                SUM(CASE WHEN status_id = 3 THEN 1 ELSE 0 END) as erledigt,
                SUM(CASE WHEN status_id = 4 THEN 1 ELSE 0 END) as abgelehnt,
                SUM(CASE WHEN status_id = 5 THEN 1 ELSE 0 END) as angebot_erstellt
            FROM lead
        """
        return self.db.fetch_one(sql)
    
    def get_lead_by_id(self, lead_id: int):
        """Gibt einen einzelnen Lead mit allen Details zurück"""
        sql = """
            SELECT 
                l.*,
                CONCAT(b_erfasser.vorname, ' ', b_erfasser.nachname) as erfasser_name,
                CONCAT(b_bearbeiter.vorname, ' ', b_bearbeiter.nachname) as bearbeiter_name,
                f.name as kunde_name,
                CONCAT(ap.vorname, ' ', ap.nachname) as ansprechpartner_name,
                ap.email as kunde_email,
                ap.telefon as kunde_telefon,
                pos.bezeichnung as ansprechpartner_position,
                p.produkt as produkt_name,
                pg.produkt as produktgruppe_name,
                s.status as status_name,
                pz.zustand as produktzustand_name,
                q.quelle as quelle_name
            FROM lead l
            LEFT JOIN benutzer b_erfasser ON l.erfasser_id = b_erfasser.benutzer_id
            LEFT JOIN benutzer b_bearbeiter ON l.bearbeiter_id = b_bearbeiter.benutzer_id
            LEFT JOIN ansprechpartner ap ON l.ansprechpartner_id = ap.id
            LEFT JOIN firma f ON ap.firma_id = f.id
            LEFT JOIN position pos ON ap.position_id = pos.id
            LEFT JOIN produkte p ON l.produkt_id = p.produkt_id
            LEFT JOIN produktgruppe pg ON l.produktgruppe_id = pg.produkt_id
            LEFT JOIN status s ON l.status_id = s.id
            LEFT JOIN produktzustand pz ON l.produktzustand_id = pz.id
            LEFT JOIN quelle q ON l.quelle_id = q.id_quelle
            WHERE l.lead_id = ?
        """
        return self.db.fetch_one(sql, (lead_id,))
    
    def get_lead_aktionen(self, lead_id: int):
        """Gibt alle Aktionen zu einem Lead zurück (Verlauf)"""
        sql = """
            SELECT 
                a.*,
                CONCAT(b.vorname, ' ', b.nachname) as benutzer_name,
                CONCAT(z.vorname, ' ', z.nachname) as ziel_name
            FROM lead_aktionen a
            LEFT JOIN benutzer b ON a.benutzer_id = b.benutzer_id
            LEFT JOIN benutzer z ON a.ziel_benutzer_id = z.benutzer_id
            WHERE a.lead_id = ?
            ORDER BY a.zeitstempel DESC
        """
        return self.db.fetch_all(sql, (lead_id,))
    
    def get_lead_kommentare(self, lead_id: int):
        """Gibt alle Kommentare zu einem Lead zurück"""
        sql = """
            SELECT *
            FROM kommentar
            WHERE lead_id = ?
            ORDER BY Datum ASC
        """
        return self.db.fetch_all(sql, (lead_id,))


# ============================================================================
# VIEW-KLASSE (UI / Darstellung)
# ============================================================================

class AuswertungView:
    """Zeigt alle Leads für Auswertungszwecke an"""
    
    def __init__(self, page: ft.Page, manager: AuswertungManager, current_user: dict, app_controller=None):
        self.page = page
        self.manager = manager
        self.current_user = current_user
        self.app_controller = app_controller
        self.all_leads = []
        self.filtered_leads = []
        self.current_filter_erfasser = None
        self.current_filter_status = None
        self.search_query = ""
        
        # UI-Referenzen
        self.erfasser_dropdown = None
        self.status_dropdown = None
        self.search_field = None
        self.lead_list_container = None
        self.stats_section = None
        
    def render(self):
        """Zeigt die Auswertungs-Ansicht"""
        self.page.clean()
        self.page.padding = 0
        
        # Leads laden
        self._load_leads()
        
        # Statistiken abrufen
        stats = self.manager.get_statistics()
        
        # Header
        header = ft.Container(
            content=ft.Row([
                ft.Row([
                    ft.IconButton(
                        icon=ft.Icons.ARROW_BACK,
                        on_click=lambda e: self._go_back(),
                        tooltip="Zurück"
                    ),
                    ft.Text("Lead Ansicht - Alle Leads", size=24, weight=ft.FontWeight.BOLD),
                ], spacing=10),
                ft.Row([
                    ft.IconButton(
                        icon=ft.Icons.DOWNLOAD,
                        icon_size=28,
                        on_click=lambda e: self._export_to_excel(),
                        tooltip="Als Excel exportieren (mit Filtern)"
                    ),
                ], spacing=15),
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
            padding=ft.padding.symmetric(horizontal=30, vertical=20),
        )
        
        # Statistik-Karten
        self.stats_section = ft.Container(
            content=ft.Row([
                self._create_stat_card("Gesamt", stats.get('gesamt', 0), "#3b82f6"),
                self._create_stat_card("Offen", stats.get('offen', 0), "#f59e0b"),
                self._create_stat_card("In Bearbeitung", stats.get('in_bearbeitung', 0), "#8b5cf6"),
                self._create_stat_card("Erledigt", stats.get('erledigt', 0), "#10b981"),
                self._create_stat_card("Abgelehnt", stats.get('abgelehnt', 0), "#ef4444"),
                self._create_stat_card("Angebot erstellt", stats.get('angebot_erstellt', 0), "#9ca3af"),
            ], spacing=15, wrap=True),
            padding=ft.padding.symmetric(horizontal=30, vertical=20),
        )
        
        # Filter-Leiste
        filter_section = self._create_filter_section()
        
        # Lead-Liste Container
        self.lead_list_container = ft.Container(
            content=self._create_lead_list(),
            expand=True,
        )
        
        # Hauptcontainer
        main_content = ft.Column([
            header,
            self.stats_section,
            filter_section,
            self.lead_list_container,
        ], spacing=0, expand=True, scroll=ft.ScrollMode.AUTO)
        
        self.page.add(main_content)
    
    def _create_stat_card(self, title, value, color):
        """Erstellt eine Statistik-Karte"""
        return ft.Container(
            content=ft.Column([
                ft.Text(title, size=14, color="#94a3b8"),
                ft.Text(str(value), size=32, weight=ft.FontWeight.BOLD),
            ], spacing=5, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            padding=20,
            border_radius=12,
            width=150,
            border=ft.border.all(2, color),
        )
    
    def _create_filter_section(self):
        """Erstellt die Filter-Leiste"""
        erfasser_list = self.manager.get_all_erfasser()
        
        self.erfasser_dropdown = ft.Dropdown(
            label="Nach Erfasser filtern",
            width=250,
            options=[ft.DropdownOption(key="all", text="Alle Erfasser")] +
                    [ft.DropdownOption(key=str(e['benutzer_id']), text=e['name']) 
                     for e in erfasser_list],
            value="all",
            on_select=lambda e: self._filter_by_erfasser(e.control.value),
        )
        
        self.status_dropdown = ft.Dropdown(
            label="Nach Status filtern",
            width=250,
            options=[
                ft.DropdownOption(key="all", text="Alle Anzeigen"),
                ft.DropdownOption(key="1", text="Offen"),
                ft.DropdownOption(key="2", text="In Bearbeitung"),
                ft.DropdownOption(key="3", text="Erledigt"),
                ft.DropdownOption(key="4", text="Abgelehnt"),
                ft.DropdownOption(key="5", text="Angebot erstellt"),
            ],
            value="all",
            on_select=lambda e: self._filter_by_status(e.control.value),
        )
        
        self.search_field = ft.TextField(
            label="Suchen",
            hint_text="Kunde, Produkt, Erfasser...",
            width=300,
            on_change=lambda e: self._search_leads(e.control.value),
            prefix_icon=ft.Icons.SEARCH,
        )
        
        reset_button = ft.ElevatedButton(
            "Filter zurücksetzen",
            icon=ft.Icons.REFRESH,
            on_click=lambda e: self._reset_filters(),
        )
        
        return ft.Container(
            content=ft.Row([
                self.erfasser_dropdown,
                self.status_dropdown,
                self.search_field,
                reset_button,
            ], spacing=15, wrap=True),
            padding=ft.padding.symmetric(horizontal=30, vertical=15),
        )
    
    def _create_lead_list(self):
        """Erstellt die Lead-Liste"""
        if not self.filtered_leads:
            return ft.Container(
                content=ft.Column([
                    ft.Icon(ft.Icons.INBOX, size=64, color="#475569"),
                    ft.Text("Keine Leads gefunden", size=18, color="#94a3b8"),
                ], horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                padding=50,
            )
        
        lead_items = []
        for lead in self.filtered_leads:
            lead_items.append(self._create_lead_card(lead))
        
        return ft.Container(
            content=ft.Column(lead_items, spacing=10, scroll=ft.ScrollMode.AUTO),
            padding=ft.padding.symmetric(horizontal=30, vertical=15),
            expand=True,
        )
    
    def _create_lead_card(self, lead):
        """Erstellt eine Lead-Karte"""
        # Farbe basierend auf Status
        status_colors = {
            "Offen": "#f59e0b",
            "In Bearbeitung": "#8b5cf6",
            "Erledigt": "#10b981",
            "Abgelehnt": "#ef4444",
        }
        status_color = status_colors.get(lead.get('status_name', 'Offen'), "#64748b")
        
        # Datum formatieren
        datum = lead.get('datum_erfasst', '')
        if isinstance(datum, datetime):
            datum_str = datum.strftime("%d.%m.%Y %H:%M")
        else:
            datum_str = str(datum)
        
        return ft.Container(
            content=ft.Column([
                # Kopfzeile
                ft.Row([
                    ft.Row([
                        ft.Container(
                            content=ft.Text(f"#{lead.get('lead_id')}", size=12, weight=ft.FontWeight.BOLD),
                            bgcolor=status_color,
                            padding=ft.padding.symmetric(horizontal=8, vertical=4),
                            border_radius=4,
                        ),
                        ft.Text(lead.get('kunde_name', 'Unbekannt'), size=16, weight=ft.FontWeight.W_600),
                    ], spacing=10),
                    ft.Container(
                        content=ft.Text(lead.get('status_name', 'Offen'), size=12, color="white"),
                        bgcolor=status_color,
                        padding=ft.padding.symmetric(horizontal=10, vertical=5),
                        border_radius=12,
                    ),
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                
                ft.Divider(height=1, color="#475569"),
                
                # Details
                ft.Row([
                    ft.Column([
                        ft.Text("Produkt:", size=12, color="#64748b"),
                        ft.Text(lead.get('produkt_name', 'Unbekannt'), size=14),
                    ], spacing=2),
                    ft.Column([
                        ft.Text("Erfasser:", size=12, color="#64748b"),
                        ft.Text(lead.get('erfasser_name', 'Unbekannt'), size=14),
                    ], spacing=2),
                    ft.Column([
                        ft.Text("Bearbeiter:", size=12, color="#64748b"),
                        ft.Text(lead.get('bearbeiter_name', 'Nicht zugewiesen'), size=14),
                    ], spacing=2),
                    ft.Column([
                        ft.Text("Erfasst am:", size=12, color="#64748b"),
                        ft.Text(datum_str, size=14),
                    ], spacing=2),
                ], spacing=30, wrap=True),
                
                # Zusätzliche Informationen
                ft.Row([
                    ft.Row([
                        ft.Icon(ft.Icons.CATEGORY_OUTLINED, size=16, color="#64748b"),
                        ft.Text(lead.get('produktgruppe_name', ''), size=12, color="#94a3b8"),
                    ], spacing=5),
                    ft.Row([
                        ft.Icon(ft.Icons.INFO_OUTLINE, size=16, color="#64748b"),
                        ft.Text(lead.get('produktzustand_name', ''), size=12, color="#94a3b8"),
                    ], spacing=5),
                    ft.Row([
                        ft.Icon(ft.Icons.SOURCE_OUTLINED, size=16, color="#64748b"),
                        ft.Text(lead.get('quelle_name', ''), size=12, color="#94a3b8"),
                    ], spacing=5),
                ], spacing=20, wrap=True),
            ], spacing=10),
            padding=20,
            border_radius=12,
            border=ft.border.all(1, ft.Colors.OUTLINE),
            on_click=lambda e, l=lead: self._show_lead_details(l),
            ink=True,
        )
    
    def _load_leads(self):
        """Lädt alle Leads aus der Datenbank"""
        self.all_leads = self.manager.get_all_leads()
        self.filtered_leads = self.all_leads.copy()
    
    def _filter_by_erfasser(self, erfasser_id):
        """Filtert Leads nach Erfasser"""
        if erfasser_id == "all":
            self.current_filter_erfasser = None
        else:
            self.current_filter_erfasser = int(erfasser_id)
        self._apply_filters()
    
    def _filter_by_status(self, status_id):
        """Filtert Leads nach Status"""
        if status_id == "all":
            self.current_filter_status = None
        else:
            self.current_filter_status = int(status_id)
        self._apply_filters()
    
    def _search_leads(self, query):
        """Sucht in Leads nach Text"""
        self.search_query = query.lower()
        self._apply_filters()
    
    def _apply_filters(self):
        """Wendet alle aktiven Filter an"""
        self.filtered_leads = self.all_leads.copy()
        
        # Erfasser-Filter
        if self.current_filter_erfasser is not None:
            self.filtered_leads = [
                lead for lead in self.filtered_leads 
                if lead.get('erfasser_id') == self.current_filter_erfasser
            ]
        
        # Status-Filter
        if self.current_filter_status is not None:
            self.filtered_leads = [
                lead for lead in self.filtered_leads 
                if lead.get('status_id') == self.current_filter_status
            ]
        
        # Suchfilter
        if self.search_query:
            self.filtered_leads = [
                lead for lead in self.filtered_leads
                if self.search_query in str(lead.get('kunde_name', '')).lower() or
                   self.search_query in str(lead.get('produkt_name', '')).lower() or
                   self.search_query in str(lead.get('erfasser_name', '')).lower() or
                   self.search_query in str(lead.get('bearbeiter_name', '')).lower()
            ]
        
        # Nur die Lead-Liste aktualisieren, nicht die gesamte Seite
        self._update_lead_list()
    
    def _reset_filters(self):
        """Setzt alle Filter zurück"""
        self.current_filter_erfasser = None
        self.current_filter_status = None
        self.search_query = ""
        if self.erfasser_dropdown:
            self.erfasser_dropdown.value = "all"
        if self.status_dropdown:
            self.status_dropdown.value = "all"
        if self.search_field:
            self.search_field.value = ""
        self._apply_filters()
    
    def _update_lead_list(self):
        """Aktualisiert nur die Lead-Liste ohne die gesamte Seite neu zu rendern"""
        if self.lead_list_container:
            self.lead_list_container.content = self._create_lead_list()
            self.page.update()
    
    def _show_lead_details(self, lead):
        """Zeigt die Detailansicht für einen Lead"""
        detail_view = LeadDetailViewAuswertung(self.page, self.manager, lead, self)
        detail_view.render()
    
    def _export_to_excel(self):
        """Exportiert ALLE Leads als Excel-Datei mit aktiven Filtern"""
        if not EXCEL_AVAILABLE:
            self._show_message(
                "Excel-Export nicht verfügbar",
                "Die benötigte Bibliothek 'openpyxl' ist nicht installiert.\n\nBitte installieren Sie: pip install openpyxl"
            )
            return
        
        if not self.all_leads:
            self._show_message("Keine Daten zum Exportieren", "Es wurden keine Leads gefunden.")
            return
        
        try:
            # Erstelle Export-Ordner falls nicht vorhanden
            export_dir = Path.home() / "Downloads"
            export_dir.mkdir(exist_ok=True)
            
            # Generiere Dateinamen mit Timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = export_dir / f"leads_export_{timestamp}.xlsx"
            
            # Excel-Arbeitsmappe erstellen
            wb = Workbook()
            ws = wb.active
            ws.title = "Leads"
            
            # Überschriften definieren
            headers = [
                'Lead-ID', 'Datum erfasst', 'Kunde', 'Ansprechpartner',
                'Produkt', 'Produktgruppe', 'Zustand', 'Quelle',
                'Status', 'Erfasser', 'Bearbeiter',
                'Angenommen am', 'Weitergeleitet am', 'Abgelehnt am', 'Erledigt am', 'Letzte Aktion'
            ]
            
            # Header-Zeile mit Formatierung
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="0f172a", end_color="0f172a", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # ALLE Leads exportieren (nicht nur gefilterte)
            for row_idx, lead in enumerate(self.all_leads, 2):
                # Datum formatieren
                datum = lead.get('datum_erfasst', '')
                if isinstance(datum, datetime):
                    datum_str = datum.strftime("%d.%m.%Y %H:%M")
                else:
                    datum_str = str(datum)
                
                # Lade Aktionen für diesen Lead
                aktionen = self.manager.get_lead_aktionen(lead.get('lead_id'))
                
                # Extrahiere spezifische Aktionsdaten
                angenommen_am = ""
                weitergeleitet_am = ""
                abgelehnt_am = ""
                erledigt_am = ""
                letzte_aktion = ""
                
                if aktionen:
                    for aktion in aktionen:
                        zeitstempel = aktion.get('zeitstempel', '')
                        if isinstance(zeitstempel, datetime):
                            zeit_str = zeitstempel.strftime("%d.%m.%Y %H:%M")
                        else:
                            zeit_str = str(zeitstempel)
                        
                        aktion_typ = aktion.get('aktion_typ', '')
                        
                        if aktion_typ == 'angenommen' and not angenommen_am:
                            angenommen_am = zeit_str
                        elif aktion_typ == 'zugewiesen' and not weitergeleitet_am:
                            weitergeleitet_am = zeit_str
                        elif aktion_typ == 'abgelehnt' and not abgelehnt_am:
                            abgelehnt_am = zeit_str
                        elif aktion_typ == 'erledigt' and not erledigt_am:
                            erledigt_am = zeit_str
                    
                    # Letzte Aktion (erste in der Liste da DESC sortiert)
                    erste_aktion = aktionen[0]
                    benutzer = erste_aktion.get('benutzer_name', 'Unbekannt')
                    typ = erste_aktion.get('aktion_typ', '')
                    letzte_aktion = f"{typ} von {benutzer}"
                
                # Daten in Zeile schreiben
                ws.cell(row=row_idx, column=1, value=lead.get('lead_id', ''))
                ws.cell(row=row_idx, column=2, value=datum_str)
                ws.cell(row=row_idx, column=3, value=lead.get('kunde_name', ''))
                ws.cell(row=row_idx, column=4, value=lead.get('ansprechpartner_name', ''))
                ws.cell(row=row_idx, column=5, value=lead.get('produkt_name', ''))
                ws.cell(row=row_idx, column=6, value=lead.get('produktgruppe_name', ''))
                ws.cell(row=row_idx, column=7, value=lead.get('produktzustand_name', ''))
                ws.cell(row=row_idx, column=8, value=lead.get('quelle_name', ''))
                ws.cell(row=row_idx, column=9, value=lead.get('status_name', ''))
                ws.cell(row=row_idx, column=10, value=lead.get('erfasser_name', ''))
                ws.cell(row=row_idx, column=11, value=lead.get('bearbeiter_name', 'Nicht zugewiesen'))
                ws.cell(row=row_idx, column=12, value=angenommen_am)
                ws.cell(row=row_idx, column=13, value=weitergeleitet_am)
                ws.cell(row=row_idx, column=14, value=abgelehnt_am)
                ws.cell(row=row_idx, column=15, value=erledigt_am)
                ws.cell(row=row_idx, column=16, value=letzte_aktion)
            
            # Spaltenbreiten anpassen
            column_widths = {
                'A': 10,  # Lead-ID
                'B': 18,  # Datum erfasst
                'C': 25,  # Kunde
                'D': 25,  # Ansprechpartner
                'E': 20,  # Produkt
                'F': 20,  # Produktgruppe
                'G': 15,  # Zustand
                'H': 15,  # Quelle
                'I': 18,  # Status
                'J': 20,  # Erfasser
                'K': 20,  # Bearbeiter
                'L': 18,  # Angenommen am
                'M': 18,  # Weitergeleitet am
                'N': 18,  # Abgelehnt am
                'O': 18,  # Erledigt am
                'P': 30,  # Letzte Aktion
            }
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # AutoFilter für alle Spalten aktivieren
            ws.auto_filter.ref = f"A1:P{len(self.all_leads) + 1}"
            
            # Aktive Filter in Excel anwenden
            filter_info = []
            
            # Status-Filter in Excel setzen
            if self.current_filter_status is not None:
                status_name = self._get_status_name(self.current_filter_status)
                if status_name:
                    filter_info.append(f"Status: {status_name}")
                    # Filter auf Spalte I (Status) - Index 8
                    # Sammle alle einzigartigen Status-Werte
                    all_status_values = set(lead.get('status_name', '') for lead in self.all_leads)
                    # Verstecke alle Zeilen die NICHT dem Filter entsprechen
                    hidden_values = [s for s in all_status_values if s != status_name]
                    if hidden_values:
                        ws.auto_filter.add_filter_column(8, hidden_values, blank=False)
            
            # Erfasser-Filter in Excel setzen
            if self.current_filter_erfasser is not None:
                erfasser_name = self._get_erfasser_name(self.current_filter_erfasser)
                if erfasser_name:
                    filter_info.append(f"Erfasser: {erfasser_name}")
                    # Filter auf Spalte J (Erfasser) - Index 9
                    # Sammle alle einzigartigen Erfasser-Werte
                    all_erfasser_values = set(lead.get('erfasser_name', '') for lead in self.all_leads)
                    # Verstecke alle Zeilen die NICHT dem Filter entsprechen
                    hidden_values = [e for e in all_erfasser_values if e != erfasser_name]
                    if hidden_values:
                        ws.auto_filter.add_filter_column(9, hidden_values, blank=False)
            
            # Suchfilter-Info (kann in Excel nicht direkt angewendet werden)
            if self.search_query:
                filter_info.append(f"Suche: '{self.search_query}' (manuell in Excel filtern)")
            
            # Zeilen ausblenden die nicht den Filtern entsprechen
            if self.current_filter_status is not None or self.current_filter_erfasser is not None:
                for row_idx, lead in enumerate(self.all_leads, 2):
                    should_hide = False
                    
                    # Prüfe Status-Filter
                    if self.current_filter_status is not None:
                        if lead.get('status_id') != self.current_filter_status:
                            should_hide = True
                    
                    # Prüfe Erfasser-Filter
                    if self.current_filter_erfasser is not None:
                        if lead.get('erfasser_id') != self.current_filter_erfasser:
                            should_hide = True
                    
                    # Zeile ausblenden
                    if should_hide:
                        ws.row_dimensions[row_idx].hidden = True
            
            # Datei speichern
            wb.save(filename)
            
            # Erfolgsmeldung mit Filterinfo
            filter_text = "\n\nAktive Filter beim Export:\n" + "\n".join(filter_info) if filter_info else ""
            self._show_message(
                "Export erfolgreich",
                #f"Es wurden {len(self.all_leads)} Leads exportiert."#\n(Aktuell gefiltert angezeigt: {len(self.filtered_leads)}){filter_text}\n
                f"\nDatei gespeichert unter:\n{filename}\n\n",#Hinweis: Alle Leads wurden exportiert.\nSie können in Excel weiter filtern.",
                success=True
            )
            
        except Exception as e:
            self._show_message(
                "Export fehlgeschlagen",
                f"Fehler beim Exportieren: {str(e)}"
            )
    
    def _get_status_name(self, status_id):
        """Gibt den Status-Namen für eine Status-ID zurück"""
        status_map = {
            1: "Offen",
            2: "In Bearbeitung",
            3: "Erledigt",
            4: "Abgelehnt",
            5: "Angebot erstellt"
        }
        return status_map.get(status_id, "")
    
    def _get_erfasser_name(self, erfasser_id):
        """Gibt den Erfasser-Namen für eine Erfasser-ID zurück"""
        if self.erfasser_dropdown and self.erfasser_dropdown.options:
            for option in self.erfasser_dropdown.options:
                if option.key == str(erfasser_id):
                    return option.text
        return ""
    
    def _show_message(self, title, message, success=False):
        """Zeigt eine Benachrichtigung an"""
        def close_dialog(e):
            self.page.pop_dialog()
        
        dialog = ft.AlertDialog(
            title=ft.Text(title),
            content=ft.Text(message),
            actions=[
                ft.TextButton("OK", on_click=close_dialog)
            ],
        )
        self.page.show_dialog(dialog)
    
    def _go_back(self):
        """Zurück zum Hauptmenü"""
        if self.app_controller:
            # Für rolle_id = 4 zurück zum Auswertungsmenü, sonst normales Hauptmenü
            if self.current_user.get('rolle_id') == 4:
                self.app_controller.show_auswertung_menu()
            else:
                self.app_controller.show_main_app()


# ============================================================================
# DETAIL-VIEW FÜR AUSWERTUNG (nur Anzeige, keine Aktionen)
# ============================================================================

class LeadDetailViewAuswertung:
    """Detailansicht eines Leads in der Auswertung (nur lesend)"""
    
    def __init__(self, page: ft.Page, manager: AuswertungManager, lead: dict, parent_view):
        self.page = page
        self.manager = manager
        self.lead = lead
        self.parent_view = parent_view
        self.aktionen = []
        self.kommentare = []
    
    def render(self):
        """Zeigt die Lead-Details"""
        self.page.clean()
        self.page.padding = 0
        
        # Aktualisierte Daten laden
        self.lead = self.manager.get_lead_by_id(self.lead.get('lead_id'))
        self.aktionen = self.manager.get_lead_aktionen(self.lead.get('lead_id'))
        self.kommentare = self.manager.get_lead_kommentare(self.lead.get('lead_id'))
        
        # Header
        header = ft.Container(
            content=ft.Row([
                ft.IconButton(
                    icon=ft.Icons.ARROW_BACK,
                    on_click=lambda e: self._go_back(),
                    tooltip="Zurück zur Liste"
                ),
                ft.Text(f"Lead #{self.lead.get('lead_id')}", size=24, weight=ft.FontWeight.BOLD),
            ], spacing=10),
            padding=ft.padding.symmetric(horizontal=30, vertical=20),
        )
        
        # Lead-Informationen
        info_section = self._build_info_section()
        
        # Verlauf / Aktionen
        verlauf_section = self._build_verlauf_section()
        
        # Kommentare
        kommentar_section = self._build_kommentar_section()
        
        # Hauptcontainer
        main_content = ft.Column([
            header,
            ft.Container(
                content=ft.Column([
                    info_section,
                    ft.Container(height=20),
                    verlauf_section,
                    ft.Container(height=20),
                    kommentar_section,
                ], scroll=ft.ScrollMode.AUTO),
                padding=ft.padding.symmetric(horizontal=30, vertical=20),
                expand=True,
            )
        ], spacing=0, expand=True)
        
        self.page.add(main_content)
    
    def _build_info_section(self):
        """Lead-Informationen"""
        # Datum formatieren
        datum = self.lead.get('datum_erfasst', '')
        if isinstance(datum, datetime):
            datum_str = datum.strftime("%d.%m.%Y %H:%M")
        else:
            datum_str = str(datum)
        
        # Status-Farbe
        status_colors = {
            "Offen": "#f59e0b",
            "In Bearbeitung": "#8b5cf6",
            "Erledigt": "#10b981",
            "Abgelehnt": "#ef4444",
        }
        status_color = status_colors.get(self.lead.get('status_name', 'Offen'), "#64748b")
        
        return ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Text("Lead-Informationen", size=20, weight=ft.FontWeight.BOLD),
                    ft.Container(
                        content=ft.Text(self.lead.get('status_name', 'Offen'), size=14),
                        bgcolor=status_color,
                        padding=ft.padding.symmetric(horizontal=15, vertical=8),
                        border_radius=20,
                    ),
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                ft.Divider(color="#475569"),
                
                # Kundendaten
                ft.Text("Kunde & Ansprechpartner", size=16, color="#94a3b8", weight=ft.FontWeight.W_600),
                ft.Row([
                    ft.Column([
                        ft.Text("Firma:", size=12, color="#64748b"),
                        ft.Text(self.lead.get('kunde_name', '-'), size=14),
                    ], spacing=2),
                    ft.Column([
                        ft.Text("Ansprechpartner:", size=12, color="#64748b"),
                        ft.Text(self.lead.get('ansprechpartner_name', '-'), size=14),
                    ], spacing=2),
                    ft.Column([
                        ft.Text("Position:", size=12, color="#64748b"),
                        ft.Text(self.lead.get('ansprechpartner_position', '-'), size=14),
                    ], spacing=2),
                ], spacing=30, wrap=True),
                ft.Row([
                    ft.Column([
                        ft.Text("E-Mail:", size=12, color="#64748b"),
                        ft.Text(self.lead.get('kunde_email', '-'), size=14),
                    ], spacing=2),
                    ft.Column([
                        ft.Text("Telefon:", size=12, color="#64748b"),
                        ft.Text(self.lead.get('kunde_telefon', '-'), size=14),
                    ], spacing=2),
                ], spacing=30, wrap=True),
                
                ft.Container(height=15),
                
                # Produktdaten
                ft.Text("Produktinformationen", size=16, color="#94a3b8", weight=ft.FontWeight.W_600),
                ft.Row([
                    ft.Column([
                        ft.Text("Produktgruppe:", size=12, color="#64748b"),
                        ft.Text(self.lead.get('produktgruppe_name', '-'), size=14),
                    ], spacing=2),
                    ft.Column([
                        ft.Text("Produkt:", size=12, color="#64748b"),
                        ft.Text(self.lead.get('produkt_name', '-'), size=14),
                    ], spacing=2),
                    ft.Column([
                        ft.Text("Zustand:", size=12, color="#64748b"),
                        ft.Text(self.lead.get('produktzustand_name', '-'), size=14),
                    ], spacing=2),
                    ft.Column([
                        ft.Text("Quelle:", size=12, color="#64748b"),
                        ft.Text(self.lead.get('quelle_name', '-'), size=14),
                    ], spacing=2),
                ], spacing=30, wrap=True),
                
                ft.Container(height=15),
                
                # Zuständigkeiten
                ft.Text("Zuständigkeiten", size=16, color="#94a3b8", weight=ft.FontWeight.W_600),
                ft.Row([
                    ft.Column([
                        ft.Text("Erfasst von:", size=12, color="#64748b"),
                        ft.Text(self.lead.get('erfasser_name', '-'), size=14),
                    ], spacing=2),
                    ft.Column([
                        ft.Text("Bearbeiter:", size=12, color="#64748b"),
                        ft.Text(self.lead.get('bearbeiter_name', 'Nicht zugewiesen'), size=14),
                    ], spacing=2),
                    ft.Column([
                        ft.Text("Erfasst am:", size=12, color="#64748b"),
                        ft.Text(datum_str, size=14),
                    ], spacing=2),
                ], spacing=30, wrap=True),
            ], spacing=10),
            padding=25,
            border_radius=12,
        )
    
    def _build_verlauf_section(self):
        """Verlauf / Aktionen"""
        if not self.aktionen:
            content = ft.Text("Keine Aktionen vorhanden", size=14, color="#64748b", italic=True)
        else:
            aktion_items = []
            for aktion in self.aktionen:
                # Zeitstempel formatieren
                zeitstempel = aktion.get('zeitstempel', '')
                if isinstance(zeitstempel, datetime):
                    zeit_str = zeitstempel.strftime("%d.%m.%Y %H:%M")
                else:
                    zeit_str = str(zeitstempel)
                
                # Aktion-Text erstellen
                aktion_typ = aktion.get('aktion_typ', '')
                benutzer = aktion.get('benutzer_name', 'Unbekannt')
                ziel = aktion.get('ziel_name', '')
                kommentar = aktion.get('kommentar', '')
                
                if aktion_typ == 'zugewiesen' and ziel:
                    text = f"{benutzer} hat den Lead an {ziel} weitergeleitet"
                elif aktion_typ == 'angenommen':
                    text = f"{benutzer} hat den Lead angenommen"
                elif aktion_typ == 'abgelehnt':
                    text = f"{benutzer} hat den Lead abgelehnt"
                elif aktion_typ == 'erledigt':
                    text = f"{benutzer} hat den Lead erledigt"
                else:
                    text = f"{benutzer}: {aktion_typ}"
                
                aktion_items.append(
                    ft.Container(
                        content=ft.Column([
                            ft.Row([
                                ft.Icon(ft.Icons.HISTORY, size=16, color="#64748b"),
                                ft.Text(zeit_str, size=12, color="#64748b"),
                            ], spacing=5),
                            ft.Text(text, size=14),
                            ft.Text(kommentar, size=12, color="#94a3b8", italic=True) if kommentar else ft.Container(),
                        ], spacing=5),
                        padding=15,
                        border_radius=8,
                        border=ft.border.all(1, ft.Colors.OUTLINE),
                    )
                )
            
            content = ft.Column(aktion_items, spacing=10)
        
        return ft.Container(
            content=ft.Column([
                ft.Text("Verlauf", size=18, weight=ft.FontWeight.BOLD),
                ft.Divider(color="#475569"),
                content,
            ], spacing=10),
            padding=25,
            border_radius=12,
        )
    
    def _build_kommentar_section(self):
        """Kommentare"""
        if not self.kommentare:
            content = ft.Text("Keine Kommentare vorhanden", size=14, color="#64748b", italic=True)
        else:
            kommentar_items = []
            for kommentar in self.kommentare:
                # Datum formatieren
                datum = kommentar.get('Datum', '')
                if isinstance(datum, datetime):
                    datum_str = datum.strftime("%d.%m.%Y %H:%M")
                else:
                    datum_str = str(datum)
                
                kommentar_items.append(
                    ft.Container(
                        content=ft.Column([
                            ft.Text(datum_str, size=12, color="#64748b"),
                            ft.Text(kommentar.get('text', ''), size=14),
                        ], spacing=5),
                        padding=15,
                        border_radius=8,
                        border=ft.border.all(1, ft.Colors.OUTLINE),
                    )
                )
            
            content = ft.Column(kommentar_items, spacing=10)
        
        return ft.Container(
            content=ft.Column([
                ft.Text("Kommentare", size=18, weight=ft.FontWeight.BOLD),
                ft.Divider(color="#475569"),
                content,
            ], spacing=10),
            padding=25,
            border_radius=12,
        )
    
    def _go_back(self):
        """Zurück zur Auswertungsliste"""
        self.parent_view.render()
