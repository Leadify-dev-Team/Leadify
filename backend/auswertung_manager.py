from backend.database import Database

# ============================================================================
# MANAGER-KLASSE (Business-Logik / Datenbankoperationen)
# ============================================================================

class AuswertungManager:
    """Verwaltet Operationen für die Auswertung aller Leads"""
    
    def __init__(self, db: Database):
        self.db = db
    
    def get_all_leads(self):
        """Gibt ALLE Leads zurück (unabhängig von Status und Bearbeiter)"""
        sql = """
            SELECT 
                l.*,
                CONCAT(b_erfasser.vorname, ' ', b_erfasser.nachname) as erfasser_name,
                CONCAT(b_bearbeiter.vorname, ' ', b_bearbeiter.nachname) as bearbeiter_name,
                f.name as kunde_name,
                CONCAT(ap.vorname, ' ', ap.nachname) as ansprechpartner_name,
                ap.email as kunde_email,
                ap.telefon as kunde_telefon,
                pos.bezeichnung as ansprechpartner_position,
                p.produkt as produkt_name,
                pg.produkt as produktgruppe_name,
                s.status as status_name,
                pz.zustand as produktzustand_name,
                q.quelle as quelle_name
            FROM lead l
            LEFT JOIN benutzer b_erfasser ON l.erfasser_id = b_erfasser.benutzer_id
            LEFT JOIN benutzer b_bearbeiter ON l.bearbeiter_id = b_bearbeiter.benutzer_id
            LEFT JOIN ansprechpartner ap ON l.ansprechpartner_id = ap.id
            LEFT JOIN firma f ON ap.firma_id = f.id
            LEFT JOIN position pos ON ap.position_id = pos.id
            LEFT JOIN produkte p ON l.produkt_id = p.produkt_id
            LEFT JOIN produktgruppe pg ON l.produktgruppe_id = pg.produkt_id
            LEFT JOIN status s ON l.status_id = s.id
            LEFT JOIN produktzustand pz ON l.produktzustand_id = pz.id
            LEFT JOIN quelle q ON l.quelle_id = q.id_quelle
            ORDER BY l.datum_erfasst DESC
        """
        return self.db.fetch_all(sql)
    
    def get_leads_by_bearbeiter(self, bearbeiter_id: int):
        """Gibt alle Leads eines bestimmten Bearbeiters zurück"""
        sql = """
            SELECT 
                l.*,
                CONCAT(b_erfasser.vorname, ' ', b_erfasser.nachname) as erfasser_name,
                CONCAT(b_bearbeiter.vorname, ' ', b_bearbeiter.nachname) as bearbeiter_name,
                f.name as kunde_name,
                CONCAT(ap.vorname, ' ', ap.nachname) as ansprechpartner_name,
                p.produkt as produkt_name,
                s.status as status_name
            FROM lead l
            LEFT JOIN benutzer b_erfasser ON l.erfasser_id = b_erfasser.benutzer_id
            LEFT JOIN benutzer b_bearbeiter ON l.bearbeiter_id = b_bearbeiter.benutzer_id
            LEFT JOIN ansprechpartner ap ON l.ansprechpartner_id = ap.id
            LEFT JOIN firma f ON ap.firma_id = f.id
            LEFT JOIN produkte p ON l.produkt_id = p.produkt_id
            LEFT JOIN status s ON l.status_id = s.id
            WHERE l.bearbeiter_id = ?
            ORDER BY l.datum_erfasst DESC
        """
        return self.db.fetch_all(sql, (bearbeiter_id,))
    
    def get_leads_by_status(self, status_id: int):
        """Gibt alle Leads mit einem bestimmten Status zurück"""
        sql = """
            SELECT 
                l.*,
                CONCAT(b_erfasser.vorname, ' ', b_erfasser.nachname) as erfasser_name,
                CONCAT(b_bearbeiter.vorname, ' ', b_bearbeiter.nachname) as bearbeiter_name,
                f.name as kunde_name,
                p.produkt as produkt_name,
                s.status as status_name
            FROM lead l
            LEFT JOIN benutzer b_erfasser ON l.erfasser_id = b_erfasser.benutzer_id
            LEFT JOIN benutzer b_bearbeiter ON l.bearbeiter_id = b_bearbeiter.benutzer_id
            LEFT JOIN ansprechpartner ap ON l.ansprechpartner_id = ap.id
            LEFT JOIN firma f ON ap.firma_id = f.id
            LEFT JOIN produkte p ON l.produkt_id = p.produkt_id
            LEFT JOIN status s ON l.status_id = s.id
            WHERE l.status_id = ?
            ORDER BY l.datum_erfasst DESC
        """
        return self.db.fetch_all(sql, (status_id,))
    
    def get_all_erfasser(self):
        """Gibt alle Erfasser zurück (für Filter)"""
        sql = """
            SELECT DISTINCT b.benutzer_id, CONCAT(b.vorname, ' ', b.nachname) as name
            FROM benutzer b
            INNER JOIN lead l ON l.erfasser_id = b.benutzer_id
            WHERE b.is_approved = 1
            ORDER BY b.vorname, b.nachname
        """
        return self.db.fetch_all(sql)
    
    def get_statistics(self):
        """Gibt Statistiken über alle Leads zurück"""
        sql = """
            SELECT 
                COUNT(*) as gesamt,
                SUM(CASE WHEN status_id = 1 THEN 1 ELSE 0 END) as offen,
                SUM(CASE WHEN status_id = 2 THEN 1 ELSE 0 END) as in_bearbeitung,
                SUM(CASE WHEN status_id = 3 THEN 1 ELSE 0 END) as erledigt,
                SUM(CASE WHEN status_id = 4 THEN 1 ELSE 0 END) as abgelehnt,
                SUM(CASE WHEN status_id = 5 THEN 1 ELSE 0 END) as angebot_erstellt
            FROM lead
        """
        return self.db.fetch_one(sql)
    
    def get_lead_by_id(self, lead_id: int):
        """Gibt einen einzelnen Lead mit allen Details zurück"""
        sql = """
            SELECT 
                l.*,
                CONCAT(b_erfasser.vorname, ' ', b_erfasser.nachname) as erfasser_name,
                CONCAT(b_bearbeiter.vorname, ' ', b_bearbeiter.nachname) as bearbeiter_name,
                f.name as kunde_name,
                CONCAT(ap.vorname, ' ', ap.nachname) as ansprechpartner_name,
                ap.email as kunde_email,
                ap.telefon as kunde_telefon,
                pos.bezeichnung as ansprechpartner_position,
                p.produkt as produkt_name,
                pg.produkt as produktgruppe_name,
                s.status as status_name,
                pz.zustand as produktzustand_name,
                q.quelle as quelle_name
            FROM lead l
            LEFT JOIN benutzer b_erfasser ON l.erfasser_id = b_erfasser.benutzer_id
            LEFT JOIN benutzer b_bearbeiter ON l.bearbeiter_id = b_bearbeiter.benutzer_id
            LEFT JOIN ansprechpartner ap ON l.ansprechpartner_id = ap.id
            LEFT JOIN firma f ON ap.firma_id = f.id
            LEFT JOIN position pos ON ap.position_id = pos.id
            LEFT JOIN produkte p ON l.produkt_id = p.produkt_id
            LEFT JOIN produktgruppe pg ON l.produktgruppe_id = pg.produkt_id
            LEFT JOIN status s ON l.status_id = s.id
            LEFT JOIN produktzustand pz ON l.produktzustand_id = pz.id
            LEFT JOIN quelle q ON l.quelle_id = q.id_quelle
            WHERE l.lead_id = ?
        """
        return self.db.fetch_one(sql, (lead_id,))
    
    def get_lead_aktionen(self, lead_id: int):
        """Gibt alle Aktionen zu einem Lead zurück (Verlauf)"""
        sql = """
            SELECT 
                a.*,
                CONCAT(b.vorname, ' ', b.nachname) as benutzer_name,
                CONCAT(z.vorname, ' ', z.nachname) as ziel_name
            FROM lead_aktionen a
            LEFT JOIN benutzer b ON a.benutzer_id = b.benutzer_id
            LEFT JOIN benutzer z ON a.ziel_benutzer_id = z.benutzer_id
            WHERE a.lead_id = ?
            ORDER BY a.zeitstempel DESC
        """
        return self.db.fetch_all(sql, (lead_id,))
    
    def get_lead_kommentare(self, lead_id: int):
        """Gibt alle Kommentare zu einem Lead zurück"""
        sql = """
            SELECT *
            FROM kommentar
            WHERE lead_id = ?
            ORDER BY Datum ASC
        """
        return self.db.fetch_all(sql, (lead_id,))
    
    # ---- Excel Export ----
    
    def export_to_excel(self, all_leads, current_filter_status=None, current_filter_erfasser=None, 
                       search_query="", erfasser_name_for_filter=""):
        """
        Exportiert Leads als Excel-Datei mit aktiven Filtern
        
        Args:
            all_leads: Liste aller Leads
            current_filter_status: Aktuelle Status-Filter-ID (optional)
            current_filter_erfasser: Aktuelle Erfasser-Filter-ID (optional)
            search_query: Aktueller Suchbegriff (optional)
            erfasser_name_for_filter: Name des Erfassers für Filter (optional)
            
        Returns:
            tuple: (success: bool, filename: Path oder None, message: str)
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from pathlib import Path
            from datetime import datetime
        except ImportError:
            return (False, None, "Die benötigte Bibliothek 'openpyxl' ist nicht installiert.\n\nBitte installieren Sie: pip install openpyxl")
        
        if not all_leads:
            return (False, None, "Es wurden keine Leads gefunden.")
        
        try:
            # Erstelle Export-Ordner falls nicht vorhanden
            export_dir = Path.home() / "Downloads"
            export_dir.mkdir(exist_ok=True)
            
            # Generiere Dateinamen mit Timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = export_dir / f"leads_export_{timestamp}.xlsx"
            
            # Excel-Arbeitsmappe erstellen
            wb = Workbook()
            ws = wb.active
            ws.title = "Leads"
            
            # Überschriften definieren
            headers = [
                'Lead-ID', 'Datum erfasst', 'Kunde', 'Ansprechpartner',
                'Produkt', 'Produktgruppe', 'Zustand', 'Quelle',
                'Status', 'Erfasser', 'Bearbeiter',
                'Angenommen am', 'Weitergeleitet am', 'Abgelehnt am', 'Erledigt am', 'Letzte Aktion'
            ]
            
            # Header-Zeile mit Formatierung
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="0f172a", end_color="0f172a", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # ALLE Leads exportieren (nicht nur gefilterte)
            for row_idx, lead in enumerate(all_leads, 2):
                # Datum formatieren
                datum = lead.get('datum_erfasst', '')
                if isinstance(datum, datetime):
                    datum_str = datum.strftime("%d.%m.%Y %H:%M")
                else:
                    datum_str = str(datum)
                
                # Lade Aktionen für diesen Lead
                aktionen = self.get_lead_aktionen(lead.get('lead_id'))
                
                # Extrahiere spezifische Aktionsdaten
                angenommen_am = ""
                weitergeleitet_am = ""
                abgelehnt_am = ""
                erledigt_am = ""
                letzte_aktion = ""
                
                if aktionen:
                    for aktion in aktionen:
                        zeitstempel = aktion.get('zeitstempel', '')
                        if isinstance(zeitstempel, datetime):
                            zeit_str = zeitstempel.strftime("%d.%m.%Y %H:%M")
                        else:
                            zeit_str = str(zeitstempel)
                        
                        aktion_typ = aktion.get('aktion_typ', '')
                        
                        if aktion_typ == 'angenommen' and not angenommen_am:
                            angenommen_am = zeit_str
                        elif aktion_typ == 'zugewiesen' and not weitergeleitet_am:
                            weitergeleitet_am = zeit_str
                        elif aktion_typ == 'abgelehnt' and not abgelehnt_am:
                            abgelehnt_am = zeit_str
                        elif aktion_typ == 'erledigt' and not erledigt_am:
                            erledigt_am = zeit_str
                    
                    # Letzte Aktion (erste in der Liste da DESC sortiert)
                    erste_aktion = aktionen[0]
                    benutzer = erste_aktion.get('benutzer_name', 'Unbekannt')
                    typ = erste_aktion.get('aktion_typ', '')
                    letzte_aktion = f"{typ} von {benutzer}"
                
                # Daten in Zeile schreiben
                ws.cell(row=row_idx, column=1, value=lead.get('lead_id', ''))
                ws.cell(row=row_idx, column=2, value=datum_str)
                ws.cell(row=row_idx, column=3, value=lead.get('kunde_name', ''))
                ws.cell(row=row_idx, column=4, value=lead.get('ansprechpartner_name', ''))
                ws.cell(row=row_idx, column=5, value=lead.get('produkt_name', ''))
                ws.cell(row=row_idx, column=6, value=lead.get('produktgruppe_name', ''))
                ws.cell(row=row_idx, column=7, value=lead.get('produktzustand_name', ''))
                ws.cell(row=row_idx, column=8, value=lead.get('quelle_name', ''))
                ws.cell(row=row_idx, column=9, value=lead.get('status_name', ''))
                ws.cell(row=row_idx, column=10, value=lead.get('erfasser_name', ''))
                ws.cell(row=row_idx, column=11, value=lead.get('bearbeiter_name', 'Nicht zugewiesen'))
                ws.cell(row=row_idx, column=12, value=angenommen_am)
                ws.cell(row=row_idx, column=13, value=weitergeleitet_am)
                ws.cell(row=row_idx, column=14, value=abgelehnt_am)
                ws.cell(row=row_idx, column=15, value=erledigt_am)
                ws.cell(row=row_idx, column=16, value=letzte_aktion)
            
            # Spaltenbreiten anpassen
            column_widths = {
                'A': 10,  # Lead-ID
                'B': 18,  # Datum erfasst
                'C': 25,  # Kunde
                'D': 25,  # Ansprechpartner
                'E': 20,  # Produkt
                'F': 20,  # Produktgruppe
                'G': 15,  # Zustand
                'H': 15,  # Quelle
                'I': 18,  # Status
                'J': 20,  # Erfasser
                'K': 20,  # Bearbeiter
                'L': 18,  # Angenommen am
                'M': 18,  # Weitergeleitet am
                'N': 18,  # Abgelehnt am
                'O': 18,  # Erledigt am
                'P': 30,  # Letzte Aktion
            }
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # AutoFilter für alle Spalten aktivieren
            ws.auto_filter.ref = f"A1:P{len(all_leads) + 1}"
            
            # Aktive Filter in Excel anwenden
            filter_info = []
            
            # Status-Filter in Excel setzen
            if current_filter_status is not None:
                status_name = self._get_status_name(current_filter_status)
                if status_name:
                    filter_info.append(f"Status: {status_name}")
                    # Filter auf Spalte I (Status) - Index 8
                    # Sammle alle einzigartigen Status-Werte
                    all_status_values = set(lead.get('status_name', '') for lead in all_leads)
                    # Verstecke alle Zeilen die NICHT dem Filter entsprechen
                    hidden_values = [s for s in all_status_values if s != status_name]
                    if hidden_values:
                        ws.auto_filter.add_filter_column(8, hidden_values, blank=False)
            
            # Erfasser-Filter in Excel setzen
            if current_filter_erfasser is not None and erfasser_name_for_filter:
                filter_info.append(f"Erfasser: {erfasser_name_for_filter}")
                # Filter auf Spalte J (Erfasser) - Index 9
                # Sammle alle einzigartigen Erfasser-Werte
                all_erfasser_values = set(lead.get('erfasser_name', '') for lead in all_leads)
                # Verstecke alle Zeilen die NICHT dem Filter entsprechen
                hidden_values = [e for e in all_erfasser_values if e != erfasser_name_for_filter]
                if hidden_values:
                    ws.auto_filter.add_filter_column(9, hidden_values, blank=False)
            
            # Suchfilter-Info (kann in Excel nicht direkt angewendet werden)
            if search_query:
                filter_info.append(f"Suche: '{search_query}' (manuell in Excel filtern)")
            
            # Zeilen ausblenden die nicht den Filtern entsprechen
            if current_filter_status is not None or current_filter_erfasser is not None:
                for row_idx, lead in enumerate(all_leads, 2):
                    should_hide = False
                    
                    # Prüfe Status-Filter
                    if current_filter_status is not None:
                        if lead.get('status_id') != current_filter_status:
                            should_hide = True
                    
                    # Prüfe Erfasser-Filter
                    if current_filter_erfasser is not None:
                        if lead.get('erfasser_id') != current_filter_erfasser:
                            should_hide = True
                    
                    # Zeile ausblenden
                    if should_hide:
                        ws.row_dimensions[row_idx].hidden = True
            
            # Datei speichern
            wb.save(filename)
            
            # Erfolgsmeldung mit Filterinfo erstellen
            filter_text = "\n\nAktive Filter beim Export:\n" + "\n".join(filter_info) if filter_info else ""
            success_message = f"\nDatei gespeichert unter:\n{filename}\n\n"
            
            return (True, filename, success_message)
            
        except Exception as e:
            return (False, None, f"Fehler beim Exportieren: {str(e)}")
    
    def _get_status_name(self, status_id):
        """Gibt den Status-Namen für eine Status-ID zurück"""
        status_map = {
            1: "Offen",
            2: "In Bearbeitung",
            3: "Erledigt",
            4: "Abgelehnt",
            5: "Angebot erstellt"
        }
        return status_map.get(status_id, "")

